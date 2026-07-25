"""Sincroniza la hoja NUEVAS CONSULTAS del Excel contra la base — el Excel gana en gestión.

A diferencia de `importar_excel.py` (que solo INSERTA y saltea lo que ya existe), este script
hace UPSERT: las consultas que ya están en la base se ACTUALIZAN con lo que diga el Excel.
Es la herramienta para el corte final: durante unos días se siguió trabajando en el Excel
(cambiando estados, cargando acciones) mientras la ingesta automática ya traía las consultas
nuevas al sistema. Esto reconcilia ambos lados.

Uso:
    DATABASE_URL=postgresql://... python sincronizar_excel.py "ruta.xlsx"            # simulacro
    DATABASE_URL=postgresql://... python sincronizar_excel.py "ruta.xlsx" --aplicar  # escribe

Sin --aplicar no toca nada: informa qué haría. Correrlo siempre así primero.

Criterio de match: (CUIT, FECHA DE RECEPCION). Las claves repetidas (misma persona consultando
dos veces el mismo día) se emparejan por orden de aparición. Las filas sin CUIT caen a
(NOMBRE, FECHA DE RECEPCION).

Acciones: se borran y reinsertan SOLO las que vinieron del Excel (`creado_por='Migración Excel'`).
Las cargadas por una persona dentro del sistema se preservan intactas — el Excel no las conoce
y borrarlas sería perder trabajo real.
"""
import sys
import unicodedata
from collections import defaultdict

import openpyxl
from sqlalchemy import text

from db import engine
from formatos import _parse_monto, _parse_fecha

HOJA = "NUEVAS CONSULTAS"
CREADO_POR_EXCEL = "Migración Excel"

# columna Excel -> campo consulta (mismo mapeo que la migración original)
MAP = {
    "FECHA DE RECEPCION": "fecha_recepcion",
    "NOMBRE/RAZON SOCIAL": "nombre",
    "CUIT": "cuit",
    "SITUACION ARCA": "situacion_arca",
    "TELEFONO": "telefono",
    "MAIL": "mail",
    "LOCALIDAD": "localidad",
    "ACTIVIDAD ECONOMICA": "actividad_economica",
    "SECTOR": "sector",
    "MONTO": "monto",
    "DESTINO DE ASISTENCIA FINANCIERA": "destino",
    "COMO SE ENTERO DE LOS CREDITOS CFI": "como_se_entero",
    "TECNICO RESPONSABLE": "tecnico",
    "DEPARTAMENTO": "departamento",
    "LOCALIDAD CONFIRMADA": "localidad_confirmada",
    "GARANTIA": "garantia",
    "LINEA": "linea",
    "PROGRAMA": "programa",
    "ARCA CONFIRMADO": "arca_confirmado",
    "MONTO CONFIRMADO": "monto_confirmado",
    "ACTIVIDAD INSCRIPTA": "actividad_inscripta",
    "SITUACION BCRA": "situacion_bcra",
    "ESTADO": "estado",
    "OBSERVACIONES": "observaciones",
    "INFORMACION EXTRA": "informacion_extra",
}

CAMPOS = list(MAP.values())


def _txt(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _norm(s):
    """Comparación tolerante para el match por nombre: sin acentos, sin dobles espacios."""
    if not s:
        return ""
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return " ".join(s.upper().split())


def _cuit(v):
    """CUIT solo dígitos — en el Excel aparece con guiones/espacios según quién lo cargó."""
    if v is None:
        return ""
    return "".join(ch for ch in str(v) if ch.isdigit())


def _clave(cuit, nombre, fecha):
    """Clave de match. Preferimos CUIT; si no hay, caemos al nombre normalizado."""
    f = str(fecha) if fecha else ""
    return ("C", _cuit(cuit), f) if _cuit(cuit) else ("N", _norm(nombre), f)


def _filas_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if HOJA not in wb.sheetnames:
        raise SystemExit(f"El archivo no tiene la hoja '{HOJA}'. Hojas: {wb.sheetnames}")
    ws = wb[HOJA]
    filas = list(ws.iter_rows(min_row=1, values_only=True))
    headers = filas[0]
    idx = {h: i for i, h in enumerate(headers) if h}
    faltan = [c for c in MAP if c not in idx]
    if faltan:
        raise SystemExit(f"Faltan columnas esperadas en la hoja: {faltan}")

    out = []
    for row in filas[1:]:
        if not any(c is not None for c in row):
            continue
        nombre = _txt(row[idx["NOMBRE/RAZON SOCIAL"]])
        cuit = _txt(row[idx["CUIT"]])
        if not nombre and not cuit:
            continue
        fecha = _parse_fecha(row[idx["FECHA DE RECEPCION"]])

        vals = {}
        for col, campo in MAP.items():
            raw = row[idx[col]]
            if campo in ("monto", "monto_confirmado"):
                vals[campo] = _parse_monto(raw)
            elif campo == "fecha_recepcion":
                vals[campo] = fecha
            elif campo == "estado":
                vals[campo] = (_txt(raw) or "CONSULTA INICIAL").upper()
            else:
                vals[campo] = _txt(raw)

        acciones = []
        for n in range(1, 11):
            ca, cf, cd = f"ACCION {n}", f"FECHA {n}", f"DETALLE {n}"
            accion = _txt(row[idx[ca]]) if ca in idx else None
            if not accion:
                continue
            acciones.append({
                "fecha": _parse_fecha(row[idx[cf]]) if cf in idx else None,
                "accion": accion.upper(),
                "detalle": (_txt(row[idx[cd]]) if cd in idx else "") or "",
            })

        out.append({"clave": _clave(cuit, nombre, fecha), "vals": vals, "acciones": acciones})
    return out


def main(path, aplicar):
    filas = _filas_excel(path)
    print(f"Excel: {len(filas)} filas con datos en '{HOJA}'\n")

    with engine.begin() as conn:
        existentes = conn.execute(text("""
            SELECT id, codigo, cuit, nombre, fecha_recepcion::date AS f FROM sde_consultas
            ORDER BY id
        """)).mappings().all()

        # índice clave -> [ids en orden], para emparejar posicionalmente los repetidos
        por_clave = defaultdict(list)
        for r in existentes:
            por_clave[_clave(r["cuit"], r["nombre"], r["f"])].append(r["id"])

        usados = defaultdict(int)
        a_actualizar, a_insertar = [], []
        for f in filas:
            ids = por_clave.get(f["clave"], [])
            n = usados[f["clave"]]
            if n < len(ids):
                usados[f["clave"]] += 1
                a_actualizar.append((ids[n], f))
            else:
                a_insertar.append(f)

        acc_excel = sum(len(f["acciones"]) for f in filas)
        acc_sistema = conn.execute(text(
            "SELECT COUNT(*) FROM sde_acciones WHERE creado_por <> :cp"), {"cp": CREADO_POR_EXCEL}
        ).scalar()

        print(f"  actualizar : {len(a_actualizar)} consultas ya existentes")
        print(f"  insertar   : {len(a_insertar)} consultas nuevas")
        print(f"  acciones   : {acc_excel} desde el Excel "
              f"(se preservan {acc_sistema} cargadas dentro del sistema)\n")

        if not aplicar:
            # Red de seguridad: una fila "nueva" cuyo CUIT ya existe con OTRA fecha casi
            # siempre es un desfasaje de fecha (zona horaria, formato), no una consulta
            # nueva — insertarla duplicaría. Se avisa para revisar antes de aplicar.
            por_cuit = defaultdict(list)
            for r in existentes:
                if _cuit(r["cuit"]):
                    por_cuit[_cuit(r["cuit"])].append((r["codigo"], r["f"]))
            sospechosas = [(f, por_cuit[_cuit(f["vals"]["cuit"])])
                           for f in a_insertar if por_cuit.get(_cuit(f["vals"]["cuit"]))]
            if sospechosas:
                print(f"  ⚠️  {len(sospechosas)} de las {len(a_insertar)} a insertar tienen un CUIT")
                print("      que YA existe en la base con otra fecha. Revisar si son consultas")
                print("      repetidas de verdad o un desfasaje de fechas:")
                for f, prev in sospechosas[:10]:
                    print(f"    - {f['vals']['nombre']} | {f['vals']['cuit']} | "
                          f"excel={f['vals']['fecha_recepcion']} | base={prev}")
                print()
            limpias = [f for f in a_insertar if not por_cuit.get(_cuit(f["vals"]["cuit"]))]
            if limpias:
                print(f"  {len(limpias)} a insertar sin CUIT previo (consultas nuevas de verdad):")
                for f in limpias[:8]:
                    print(f"    - {f['vals']['nombre']} | {f['vals']['cuit']} | {f['vals']['fecha_recepcion']}")
            print("\nSIMULACRO — no se escribió nada. Volvé a correr con --aplicar para confirmar.")
            conn.rollback()
            return

        sets = ", ".join(f"{c} = :{c}" for c in CAMPOS)
        for cid, f in a_actualizar:
            conn.execute(text(f"UPDATE sde_consultas SET {sets}, updated_at = NOW() WHERE id = :id"),
                         {**f["vals"], "id": cid})

        for f in a_insertar:
            n = conn.execute(text("SELECT nextval('sde_consultas_codigo_seq')")).scalar()
            cols = ", ".join(["codigo", "fuente"] + CAMPOS)
            binds = ", ".join([":codigo", ":fuente"] + [f":{c}" for c in CAMPOS])
            cid = conn.execute(
                text(f"INSERT INTO sde_consultas ({cols}) VALUES ({binds}) RETURNING id"),
                {**f["vals"], "codigo": f"SDE-{n:06d}", "fuente": "Histórico"},
            ).scalar()
            f["_id"] = cid

        # acciones: reemplazar solo las que vinieron del Excel, preservar las del sistema
        ids_tocados = [cid for cid, _ in a_actualizar] + [f["_id"] for f in a_insertar]
        conn.execute(text("""
            DELETE FROM sde_acciones WHERE creado_por = :cp AND consulta_id = ANY(:ids)
        """), {"cp": CREADO_POR_EXCEL, "ids": ids_tocados})

        n_acc = 0
        for cid, f in a_actualizar:
            for a in f["acciones"]:
                conn.execute(text("""
                    INSERT INTO sde_acciones (consulta_id, fecha, accion, detalle, creado_por)
                    VALUES (:cid, :fecha, :accion, :detalle, :cp)
                """), {**a, "cid": cid, "cp": CREADO_POR_EXCEL})
                n_acc += 1
        for f in a_insertar:
            for a in f["acciones"]:
                conn.execute(text("""
                    INSERT INTO sde_acciones (consulta_id, fecha, accion, detalle, creado_por)
                    VALUES (:cid, :fecha, :accion, :detalle, :cp)
                """), {**a, "cid": f["_id"], "cp": CREADO_POR_EXCEL})
                n_acc += 1

        print(f"✓ Aplicado: {len(a_actualizar)} actualizadas | {len(a_insertar)} insertadas "
              f"| {n_acc} acciones del Excel")


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if not args:
        print(__doc__)
        sys.exit(1)
    main(args[0], "--aplicar" in sys.argv)
