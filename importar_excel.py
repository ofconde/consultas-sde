"""Migración inicial (RUN-ONCE): hoja NUEVAS CONSULTAS del Excel → sde_consultas + sde_acciones.

Uso:
    DATABASE_URL=postgresql://... python importar_excel.py "ruta/al/FORMULARIO ... .xlsx"

Idempotente: dedupe por (cuit + fecha_recepcion). Re-correr no duplica.
Las consultas quedan con fuente='Histórico'. Las acciones se explotan de FECHA/ACCION/DETALLE 1..10.
"""
import sys
import openpyxl
from sqlalchemy import text

from db import engine, init_db
from formatos import _parse_monto, _parse_fecha

HOJA = "NUEVAS CONSULTAS"

# columna Excel -> campo consulta
MAP = {
    "FECHA DE RECEPCION": "fecha_recepcion",
    "NOMBRE/RAZON SOCIAL": "nombre",
    "CUIT": "cuit",
    "SITUACION ARCA": "situacion_arca",
    "TELEFONO": "telefono",
    "MAIL": "mail",
    "LOCALIDAD": "localidad",
    "ACTIVIDAD ECONOMICA": "actividad_economica",
    "SECTOR": "sector",
    "MONTO": "monto",
    "DESTINO DE ASISTENCIA FINANCIERA": "destino",
    "COMO SE ENTERO DE LOS CREDITOS CFI": "como_se_entero",
    "TECNICO RESPONSABLE": "tecnico",
    "DEPARTAMENTO": "departamento",
    "LOCALIDAD CONFIRMADA": "localidad_confirmada",
    "GARANTIA": "garantia",
    "LINEA": "linea",
    "PROGRAMA": "programa",
    "ARCA CONFIRMADO": "arca_confirmado",
    "MONTO CONFIRMADO": "monto_confirmado",
    "ACTIVIDAD INSCRIPTA": "actividad_inscripta",
    "SITUACION BCRA": "situacion_bcra",
    "ESTADO": "estado",
    "OBSERVACIONES": "observaciones",
    "INFORMACION EXTRA": "informacion_extra",
}


def _txt(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def main(path):
    init_db()
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[HOJA]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers) if h}

    insertadas = duplicadas = acc_total = 0

    with engine.begin() as conn:
        # contador local de códigos (no depende de COUNT dentro de la transacción)
        seq = conn.execute(text("SELECT COUNT(*) FROM sde_consultas")).scalar() or 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(c is not None for c in row):
                continue
            nombre = _txt(row[idx["NOMBRE/RAZON SOCIAL"]]) if "NOMBRE/RAZON SOCIAL" in idx else None
            cuit = _txt(row[idx["CUIT"]]) if "CUIT" in idx else None
            if not nombre and not cuit:
                continue

            fecha_rec = _parse_fecha(row[idx["FECHA DE RECEPCION"]]) if "FECHA DE RECEPCION" in idx else None

            # dedupe por cuit + fecha_recepcion
            if cuit and fecha_rec:
                dup = conn.execute(text("""
                    SELECT id FROM sde_consultas
                    WHERE cuit = :c AND fecha_recepcion::date = :f AND fuente = 'Histórico'
                    LIMIT 1
                """), {"c": cuit, "f": fecha_rec}).scalar()
                if dup:
                    duplicadas += 1
                    continue

            seq += 1
            vals = {"codigo": f"SDE-{seq:06d}", "fuente": "Histórico"}
            for col, campo in MAP.items():
                if col not in idx:
                    continue
                raw = row[idx[col]]
                if campo in ("monto", "monto_confirmado"):
                    vals[campo] = _parse_monto(raw)
                elif campo == "fecha_recepcion":
                    vals[campo] = fecha_rec
                elif campo == "estado":
                    vals[campo] = (_txt(raw) or "CONSULTA INICIAL").upper()
                else:
                    vals[campo] = _txt(raw)

            cols = ", ".join(vals.keys())
            binds = ", ".join(f":{k}" for k in vals.keys())
            cid = conn.execute(
                text(f"INSERT INTO sde_consultas ({cols}) VALUES ({binds}) RETURNING id"),
                vals,
            ).scalar()
            insertadas += 1

            # acciones 1..10
            for n in range(1, 11):
                ca, cf, cd = f"ACCION {n}", f"FECHA {n}", f"DETALLE {n}"
                accion = _txt(row[idx[ca]]) if ca in idx else None
                if not accion:
                    continue
                conn.execute(text("""
                    INSERT INTO sde_acciones (consulta_id, fecha, accion, detalle, creado_por)
                    VALUES (:cid, :fecha, :accion, :detalle, 'Migración Excel')
                """), {
                    "cid": cid,
                    "fecha": _parse_fecha(row[idx[cf]]) if cf in idx else None,
                    "accion": accion.upper(),
                    "detalle": (_txt(row[idx[cd]]) if cd in idx else "") or "",
                })
                acc_total += 1

    print(f"✓ Importadas: {insertadas} consultas | {acc_total} acciones | duplicadas saltadas: {duplicadas}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python importar_excel.py <ruta_al_xlsx>")
        sys.exit(1)
    main(sys.argv[1])
